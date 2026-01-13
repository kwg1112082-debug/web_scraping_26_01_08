import requests
from bs4 import BeautifulSoup
import pandas as pd  # [추가] 엑셀 저장을 위한 pandas 라이브러리 (별칭: pd)

# ==========================================================
# [설정] 스크래핑할 대상 URL 및 옵션
# ==========================================================
TARGET_URL = "https://news.naver.com/section/104"
MAX_NEWS_LIMIT = 50  # 최대 수집할 뉴스 개수
EXCEL_FILENAME = "naver_news_20260113.xlsx" # [추가] 저장할 엑셀 파일 이름

def main():
    # 1단계: 웹페이지 요청 (Requests) - [ver.3 기능: 안전성 강화]
    print(f"\n[접속 시도] {TARGET_URL} 데이터를 가져오는 중입니다...")
    
    try:
        # timeout=5: 5초 안에 응답이 없으면 에러 발생 (무한 대기 방지)
        response = requests.get(TARGET_URL, timeout=5)
        
        # HTTP 상태 코드 검사
        if response.status_code != 200:
            print(f"❌ 접속 실패! 상태 코드: {response.status_code}")
            return
            
    except Exception as e:
        print(f"❌ 접속 중 오류 발생: {e}")
        return

    print("✅ 접속 성공! HTML 분석을 시작합니다.\n")


    # 2단계: HTML 파싱 (BeautifulSoup)
    soup = BeautifulSoup(response.text, "lxml")


    # 3단계: 뉴스 데이터 추출
    
    # [ver.3 기능: 섹션 제목 추출]
    section_title_tag = soup.select_one(".sa_head_link")
    if section_title_tag:
        section_name = section_title_tag.get_text(strip=True)
        print(f"📌 현재 섹션: {section_name}")
    else:
        print("📌 섹션 정보를 찾을 수 없습니다.")

    # 네이버 뉴스 목록의 각 기사는 <div class="sa_text"> 안에 담겨 있습니다.
    articles = soup.select('.sa_text')
    
    # 수집 개수 제한 (slicing 사용)
    articles = articles[:MAX_NEWS_LIMIT]
    
    print(f"[{MAX_NEWS_LIMIT}개 제한] 탐색된 기사 요소 개수: {len(articles)}개")
    print("=" * 60)

    # 수집된 데이터를 저장할 리스트
    news_data_list = []

    for article in articles:
        # (1) 제목 및 링크 추출 using CSS Selector
        # <a class="sa_text_title" href="..."> ... </a>
        title_tag = article.select_one('.sa_text_title')
        
        # 제목 태그가 없는 경우(광고 등)는 건너뜀
        if not title_tag:
            continue
            
        # 제목 텍스트 (strong 태그 안)
        strong_tag = title_tag.select_one('.sa_text_strong')
        title_text = strong_tag.get_text(strip=True) if strong_tag else title_tag.get_text(strip=True)
        
        # 기사 링크 (href 속성)
        article_link = title_tag['href']

        # (2) 기사 요약 내용 추출
        # 목록에서 보이는 짤막한 내용 (sa_text_lede)
        lede_tag = article.select_one('.sa_text_lede')
        content_text = lede_tag.get_text(strip=True) if lede_tag else "내용 미리보기 없음"

        # (3) 언론사 정보 추출
        press_tag = article.select_one('.sa_text_press')
        press_name = press_tag.get_text(strip=True) if press_tag else "언론사 정보 없음"

        # [요청 사항 반영 시작]
        # 1. 내용(content)을 15글자 이내로 제한 (넘치면 '...' 추가)
        limit_content = content_text[:15] + "..." if len(content_text) > 15 else content_text

        # 2. 제목(title)에서 '중요한 글자'만 추출 (간이 키워드 추출)
        import re
        # 제목에서 [속보], [단독] 같은 대괄호 꼬리표 제거
        clean_title = re.sub(r'\[.*?\]', '', title_text).strip()
        # 공백 기준 앞의 핵심 단어 4개만 뽑아서 '핵심 제목' 생성
        title_keywords = " ".join(clean_title.split()[:4])
        # [요청 사항 반영 끝]

        # 딕셔너리로 묶어서 리스트에 추가
        news_info = {
            "press": press_name,
            "important_title": title_keywords, # 요약된 제목
            "content": limit_content,          # 15자 제한 내용
            "link": article_link,
            "full_title": title_text           # 원본 제목 (대조용)
        }
        news_data_list.append(news_info)


    # 4단계: 결과 출력 화면에 보여주기
    for idx, news in enumerate(news_data_list, start=1):
        print(f"📰 No.{idx}")
        print(f"   신문사: {news['press']}")
        print(f"   핵심 제목: {news['important_title']}")
        print(f"   요약 내용: {news['content']}")
        print(f"   링크: {news['link']}")
        print("-" * 60)

    print(f"\n✅ 최종 수집 완료: 총 {len(news_data_list)}개의 뉴스를 정리했습니다.")


    # 5단계: 엑셀 파일로 저장하기 (Pandas + openpyxl 스타일링)
    if len(news_data_list) > 0:
        print(f"\n[엑셀 저장] '{EXCEL_FILENAME}' 파일로 저장을 시도합니다...")
        
        # 1. 데이터프레임 생성
        df = pd.DataFrame(news_data_list)
        
        # 2. 스타일링을 위해 ExcelWriter 연기
        # 'openpyxl' 엔진을 사용하여 상세한 스타일 설정을 진행합니다.
        with pd.ExcelWriter(EXCEL_FILENAME, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='네이버뉴스')
            
            # openpyxl 객체 가져오기
            workbook  = writer.book
            worksheet = writer.sheets['네이버뉴스']
            
            # 스타일을 위한 도구 임포트
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # 스타일 정의
            header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid') # 연녹색 배경
            header_font = Font(bold=True, size=12)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            highlight_font = Font(color='FF0000', bold=True) # AI 포함 시 빨간색 굵게

            # 3. 제목 행(첫 번째 행) 스타일 적용
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # 4. 데이터 영역 스타일 적용 및 'AI' 강조 필터링
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=1):
                # 제목 컬럼(중요한 제목 또는 전체 제목)에서 'AI' 검색
                # (두 번째 컬럼인 important_title과 다섯 번째 컬럼인 full_title을 모두 확인)
                row_title = str(row[1].value) + str(row[4].value) 
                is_ai_included = 'ai' in row_title.lower()

                for cell in row:
                    cell.border = thin_border
                    cell.alignment = left_alignment
                    # 만약 제목에 'ai'가 포함되어 있다면 해당 행 강조
                    if is_ai_included:
                        cell.font = highlight_font

            # 5. 열 너비 자동 조정 (내용을 한눈에 보기 쉽게)
            # 글자 수에 비례하여 너비를 넓힙니다.
            column_widths = {
                'A': 15, # 신문사
                'B': 40, # 핵심 제목
                'C': 35, # 요약 내용
                'D': 60, # 링크
                'E': 70  # 원본 제목
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

        print(f"🎉 스타일링 완료! '{EXCEL_FILENAME}' 파일이 생성되었습니다.")
    else:
        print("⚠️ 저장할 데이터가 없어 엑셀 파일을 생성하지 않았습니다.")

# 스크립트 실행
if __name__ == "__main__":
    main()
